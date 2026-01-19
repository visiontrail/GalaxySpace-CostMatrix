"""
Excel 数据处理服务
负责读取、分析、处理 Excel 文件
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Any, Optional
from datetime import datetime
import re
import os
import time
from collections import Counter

from app.utils.logger import get_logger


class ExcelProcessor:
    """Excel 处理器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.sheets_data: Dict[str, pd.DataFrame] = {}
        self.workbook = None
        self.logger = get_logger("excel_processor")
        self._attendance_cache: Optional[pd.DataFrame] = None
        self._travel_cache: Dict[str, pd.DataFrame] = {}
        self._combined_travel_cache: Optional[pd.DataFrame] = None
        
    def load_all_sheets(self, load_workbook_obj: bool = False) -> Dict[str, pd.DataFrame]:
        """
        加载所有 Sheet 数据

        Args:
            load_workbook_obj: 是否同时加载 openpyxl Workbook 对象（仅在需要回写时启用）
        """
        try:
            start = time.perf_counter()
            self.logger.info(f"开始读取 Excel 文件: {self.file_path}")
            all_sheets = pd.read_excel(self.file_path, sheet_name=None)
            elapsed = time.perf_counter() - start

            self.sheets_data = all_sheets
            self._attendance_cache = None
            self._travel_cache = {}
            self._combined_travel_cache = None

            sheet_names = ", ".join(all_sheets.keys())
            self.logger.info(f"Excel 读取完成（{sheet_names}），耗时 {elapsed:.2f}s")
            
            # 输出每个 Sheet 的基本信息
            for sheet_name, df in all_sheets.items():
                self.logger.info(f"Sheet [{sheet_name}]: {len(df)} 行, {len(df.columns)} 列")
                self.logger.info(f"  列名: {list(df.columns)}")
                if len(df) > 0:
                    self.logger.info(f"  前2行数据预览:")
                    for idx in range(min(2, len(df))):
                        row_data = df.iloc[idx].to_dict()
                        self.logger.info(f"    行{idx}: {row_data}")
            
            # 部分分析场景不需要 Workbook，仅在回写等场景按需加载
            if load_workbook_obj:
                wb_start = time.perf_counter()
                self.workbook = load_workbook(self.file_path, keep_links=False)
                self.logger.info(f"openpyxl 工作簿加载完成，耗时 {time.perf_counter() - wb_start:.2f}s")
            else:
                self.workbook = None
            
            return self.sheets_data
        except Exception as e:
            raise Exception(f"读取 Excel 文件失败: {str(e)}")

    def get_sheet_names(self) -> List[str]:
        """
        仅获取 Sheet 名称，避免读取全部数据导致耗时
        """
        try:
            workbook = load_workbook(
                self.file_path,
                read_only=True,
                data_only=True,
                keep_links=False
            )
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            raise Exception(f"读取 Excel Sheet 名称失败: {str(e)}")
    
    def get_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """获取指定 Sheet"""
        return self.sheets_data.get(sheet_name)
    
    def clean_attendance_data(self, use_cache: bool = True) -> pd.DataFrame:
        """
        清洗考勤数据（状态明细）
        """
        if use_cache and self._attendance_cache is not None:
            return self._attendance_cache

        df = self.get_sheet("状态明细")
        if df is None:
            return pd.DataFrame()
        
        # 标准化列名
        df = df.copy()
        
        # 处理日期格式
        if '日期' in df.columns:
            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
        
        # 处理工时数据
        if '工时' in df.columns:
            df['工时'] = pd.to_numeric(df['工时'], errors='coerce')
        
        # 删除空行
        df = df.dropna(subset=['姓名'], how='all')

        if use_cache:
            self._attendance_cache = df
        
        return df
    
    def clean_travel_data(self, sheet_name: str, use_cache: bool = True) -> pd.DataFrame:
        """
        清洗差旅数据（机票/酒店/火车票）
        """
        if use_cache and sheet_name in self._travel_cache:
            return self._travel_cache[sheet_name]

        df = self.get_sheet(sheet_name)
        if df is None:
            self.logger.warning(f"[{sheet_name}] Sheet 不存在")
            return pd.DataFrame()

        self.logger.info(f"[{sheet_name}] 开始清洗数据 - 原始列名: {list(df.columns)}")
        self.logger.info(f"[{sheet_name}] 原始行数: {len(df)}")

        df = df.copy()
        # 标准化列名：去除首尾空格，避免不同月份 Excel 列名细微差异导致匹配失败
        df.columns = [str(c).strip() for c in df.columns]
        original_df = df.copy()
        
        # 处理金额字段
        amount_col = '授信金额' if '授信金额' in df.columns else '金额'
        self.logger.info(f"[{sheet_name}] 金额列: {amount_col}")
        if amount_col in df.columns:
            df[amount_col] = pd.to_numeric(df[amount_col], errors='coerce')
            # 将 NaN 填充为 0，但保留所有记录
            df[amount_col] = df[amount_col].fillna(0)
            self.logger.info(f"[{sheet_name}] 金额列有效值数: {df[amount_col].notna().sum()}")
        else:
            self.logger.warning(f"[{sheet_name}] 未找到金额列（授信金额或金额）")
        
        def _parse_datetime_avoiding_time_only(series: pd.Series) -> pd.Series:
            """
            将列解析为 datetime，并避免把纯时间值（如 '22:17' 或 datetime.time）解析为“今天”的日期。
            """
            if series is None:
                return pd.Series(dtype="datetime64[ns]")

            # 已经是 datetime 类型
            try:
                if pd.api.types.is_datetime64_any_dtype(series):
                    return pd.to_datetime(series, errors="coerce")
            except Exception:
                pass

            # 处理 datetime.time 或纯时间字符串
            time_only_regex = re.compile(r"^\\s*\\d{1,2}:\\d{2}(:\\d{2})?\\s*$")

            def _is_time_obj(v: Any) -> bool:
                try:
                    from datetime import time as dt_time
                    return isinstance(v, dt_time)
                except Exception:
                    return False

            if series.dtype == object:
                cleaned = series.copy()
                mask_time_obj = cleaned.map(_is_time_obj)
                cleaned.loc[mask_time_obj] = None

                str_series = cleaned.astype(str)
                mask_time_str = str_series.str.match(time_only_regex, na=False)
                cleaned.loc[mask_time_str] = None

                return pd.to_datetime(cleaned, errors="coerce")

            return pd.to_datetime(series, errors="coerce")

        found_date_cols: List[str] = []
        if sheet_name == '机票':
            if '起飞时间' in original_df.columns:
                df['起飞日期'] = _parse_datetime_avoiding_time_only(original_df['起飞时间'])
                found_date_cols.append('起飞时间→起飞日期')
            if '起飞时间.1' in original_df.columns:
                df['起飞日期.1'] = _parse_datetime_avoiding_time_only(original_df['起飞时间.1'])
                found_date_cols.append('起飞时间.1→起飞日期.1')
        elif sheet_name == '酒店':
            # 以“入住日期”为主；若存在“入住时间”（含日期时间），补充到“入住日期.1”
            if '入住日期' in original_df.columns:
                df['入住日期'] = _parse_datetime_avoiding_time_only(original_df['入住日期'])
                found_date_cols.append('入住日期')
            if '入住时间' in original_df.columns:
                dt_full = _parse_datetime_avoiding_time_only(original_df['入住时间'])
                if dt_full.notna().any():
                    df['入住日期.1'] = dt_full
                    found_date_cols.append('入住时间→入住日期.1')
        elif sheet_name == '火车票':
            # “出发日期”是关键日期字段。旧逻辑会把“出发时间”(HH:MM)写入“出发日期”，导致日期被解析成“今天”。
            if '出发日期' in original_df.columns:
                df['出发日期'] = _parse_datetime_avoiding_time_only(original_df['出发日期'])
                found_date_cols.append('出发日期')
            elif '出发时间' in original_df.columns:
                # 兜底：某些模板可能只提供“出发时间”(包含日期时间)
                df['出发日期'] = _parse_datetime_avoiding_time_only(original_df['出发时间'])
                found_date_cols.append('出发时间→出发日期')

            # 如果“出发时间”存在且是完整日期时间，写入“出发日期.1”；若仅是时间字符串，则与“出发日期”组合
            if '出发时间' in original_df.columns:
                dt_full = _parse_datetime_avoiding_time_only(original_df['出发时间'])
                if dt_full.notna().any():
                    df['出发日期.1'] = dt_full
                    found_date_cols.append('出发时间→出发日期.1')
                elif '出发日期' in df.columns and df['出发日期'].notna().any():
                    time_str = original_df['出发时间'].astype(str).str.strip()
                    time_only_mask = time_str.str.match(r"^\\d{1,2}:\\d{2}(:\\d{2})?$", na=False)
                    if time_only_mask.any():
                        date_str = df['出发日期'].dt.strftime('%Y-%m-%d')
                        combined = pd.to_datetime(date_str + ' ' + time_str, errors='coerce')
                        df.loc[time_only_mask, '出发日期.1'] = combined.loc[time_only_mask]
                        found_date_cols.append('出发日期+出发时间→出发日期.1')

        self.logger.info(f"[{sheet_name}] 找到的日期列: {found_date_cols}")
        
        # 统一差旅人员姓名字段
        name_cols = [col for col in ['差旅人员姓名', '预订人姓名'] if col in df.columns]
        self.logger.info(f"[{sheet_name}] 找到的姓名列: {name_cols}")
        if '差旅人员姓名' in df.columns:
            df['姓名'] = df['差旅人员姓名']
        elif '预订人姓名' in df.columns:
            df['姓名'] = df['预订人姓名']
        else:
            self.logger.warning(f"[{sheet_name}] 未找到姓名列（差旅人员姓名或预订人姓名）")

        self.logger.info(f"[{sheet_name}] 清洗后行数: {len(df)}")
        self.logger.info(f"[{sheet_name}] 最终列名: {list(df.columns)}")

        if use_cache:
            self._travel_cache[sheet_name] = df
        
        return df

    def _get_combined_travel_df(self) -> pd.DataFrame:
        """
        获取带消费日期和差旅类型的汇总差旅数据（使用缓存避免重复计算）
        """
        if self._combined_travel_cache is not None:
            return self._combined_travel_cache

        frames: List[pd.DataFrame] = []
        date_columns = {
            '机票': ['起飞日期', '起飞日期.1', '起飞时间', '起飞时间.1'],
            '酒店': ['入住日期', '入住时间'],
            '火车票': ['出发日期', '出发时间']
        }

        for sheet_name, date_cols in date_columns.items():
            df = self.clean_travel_data(sheet_name)
            if df.empty:
                continue
            
            date_col = None
            for col in date_cols:
                if col in df.columns:
                    date_col = col
                    break
            
            if not date_col:
                continue

            temp = df[['姓名', date_col]].copy()
            temp = temp[temp[date_col].notna()]
            if temp.empty:
                continue

            temp['消费日期'] = temp[date_col].dt.date
            temp['差旅类型'] = sheet_name
            frames.append(temp[['姓名', '消费日期', '差旅类型']])

        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
            columns=['姓名', '消费日期', '差旅类型']
        )
        self._combined_travel_cache = combined
        return combined
    
    def extract_project_code(self, project_str: str) -> Tuple[str, str]:
        """
        从项目字段提取项目代码和名称
        格式: "05010013 市场-整星..."
        """
        if pd.isna(project_str) or not isinstance(project_str, str):
            return "", ""
        
        # 尝试提取项目代码（通常是开头的数字）
        match = re.match(r'(\d+)\s+(.*)', project_str.strip())
        if match:
            return match.group(1), match.group(2)
        
        return "", project_str
    
    def aggregate_project_costs(self, top_n: int = 20) -> List[Dict[str, Any]]:
        """
        项目成本归集
        
        Args:
            top_n: 返回前N个项目，其余汇总到"其他"（默认20）
        """
        self.logger.info("=" * 80)
        self.logger.info("开始执行项目成本归集 - 详细模式（Backend服务）")
        self.logger.info("=" * 80)
        
        results = []
        
        # 处理所有差旅相关的 Sheet
        travel_sheets = ['机票', '酒店', '火车票']
        
        all_records = []
        sheet_stats = {}
        
        for sheet_name in travel_sheets:
            self.logger.info(f"\n📋 处理差旅表: {sheet_name}")
            
            # 获取原始数据（未清洗）以获取真实行数
            df_raw = self.get_sheet(sheet_name)
            original_count = 0 if df_raw is None else len(df_raw)
            
            df = self.clean_travel_data(sheet_name)
            if df.empty:
                self.logger.warning(f"   ⚠️  {sheet_name} 数据为空")
                continue
            
            # 检查是否有项目字段
            if '项目' not in df.columns:
                self.logger.warning(f"   ⚠️  {sheet_name} 缺少'项目'列")
                continue
            
            amount_col = '授信金额' if '授信金额' in df.columns else '金额'
            date_cols = ['出发日期', '出发日期.1', '出发时间', '起飞日期', '起飞日期.1', '起飞时间', '起飞时间.1', '入住日期', '入住时间']
            date_col = next((col for col in date_cols if col in df.columns), None)
            
            self.logger.info(f"   - 原始记录数: {original_count}")
            self.logger.info(f"   - 清洗后记录数: {len(df)}")
            self.logger.info(f"   - 金额列: {amount_col}")
            
            # 统计信息
            record_count = 0
            empty_project_count = 0
            sheet_total_amount = 0
            
            for idx, row in df.iterrows():
                project_str = row.get('项目', '')
                project_code, project_name = self.extract_project_code(project_str)
                amount = row.get(amount_col, 0)
                
                # 空项目作为单独的项目类别处理
                if not project_code:
                    project_code = '空项目'
                    project_name = '未分配项目'
                    empty_project_count += 1
                
                all_records.append({
                    'project_code': project_code,
                    'project_name': project_name,
                    'amount': amount,
                    'type': sheet_name,
                    'person': row.get('姓名', ''),
                    'date': row.get(date_col, '') if date_col else ''
                })
                record_count += 1
                sheet_total_amount += amount
                
                # 输出前3条记录的详细信息
                if record_count <= 3:
                    person = row.get('姓名', '未知')
                    date_val = row.get(date_col, '') if date_col else ''
                    # 安全的日期格式化
                    if pd.notna(date_val) and hasattr(date_val, 'strftime'):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val) if pd.notna(date_val) else '未知'
                    self.logger.debug(f"      记录{record_count}: {project_code} | {person} | ¥{amount:,.2f} | {date_str}")
            
            sheet_stats[sheet_name] = {
                'original_total': original_count,
                'cleaned_total': len(df),
                'record_count': record_count,
                'empty_project_count': empty_project_count,
                'amount': sheet_total_amount
            }
            
            # 统计金额分布 - 基于所有清洗后记录
            zero_amount_count = (df[amount_col] == 0).sum() if amount_col in df.columns else 0
            negative_amount_count = (df[amount_col] < 0).sum() if amount_col in df.columns else 0
            positive_amount_count = (df[amount_col] > 0).sum() if amount_col in df.columns else 0
            filtered_count = original_count - len(df)
            
            # 计算正数/负数金额总和（基于所有清洗后记录）
            negative_amount_sum = df[df[amount_col] < 0][amount_col].sum() if amount_col in df.columns and negative_amount_count > 0 else 0
            positive_amount_sum = df[df[amount_col] > 0][amount_col].sum() if amount_col in df.columns and positive_amount_count > 0 else 0
            # 所有记录的净总金额（用于日志显示，确保与正数+负数一致）
            sheet_all_amount = df[amount_col].sum() if amount_col in df.columns else 0
            
            self.logger.info(f"   ✅ 处理完成:")
            self.logger.info(f"      - 总记录数: {record_count}")
            if empty_project_count > 0:
                self.logger.info(f"      - 空项目记录: {empty_project_count} (已归入\"空项目\"类别)")
            if filtered_count > 0:
                self.logger.warning(f"      ⚠️  数据清洗时过滤记录: {filtered_count}条")
            
            # 金额分布统计（基于所有清洗后记录）
            self.logger.info(f"      - 金额分布（全部记录）:")
            if positive_amount_count > 0:
                self.logger.info(f"         • 正数金额: {positive_amount_count}条, 合计 ¥{positive_amount_sum:,.2f}")
            if negative_amount_count > 0:
                self.logger.warning(f"         • 负数金额: {negative_amount_count}条, 合计 ¥{negative_amount_sum:,.2f} (退款/调整)")
            if zero_amount_count > 0:
                self.logger.info(f"         • 零值金额: {zero_amount_count}条")
            self.logger.info(f"         • 净总金额: ¥{sheet_all_amount:,.2f}")
        
        # 输出汇总统计
        self.logger.info(f"\n📊 差旅数据汇总:")
        original_total_records = sum(stats['original_total'] for stats in sheet_stats.values())
        cleaned_total_records = sum(stats['cleaned_total'] for stats in sheet_stats.values())
        total_records = sum(stats['record_count'] for stats in sheet_stats.values())
        empty_project_records = sum(stats['empty_project_count'] for stats in sheet_stats.values())
        total_amount = sum(stats['amount'] for stats in sheet_stats.values())
        
        self.logger.info(f"   - 原始总记录数: {original_total_records}")
        self.logger.info(f"   - 清洗后总记录数: {cleaned_total_records}")
        self.logger.info(f"   - 处理记录数: {total_records}")
        if empty_project_records > 0:
            self.logger.info(f"   - 空项目记录数: {empty_project_records} (已归入\"空项目\"类别)")
        if original_total_records > cleaned_total_records:
            filtered = original_total_records - cleaned_total_records
            self.logger.warning(f"   ⚠️  数据清洗过滤了 {filtered} 条记录（可能是无效数据或删除行）")
        self.logger.info(f"   - 净总金额: ¥{total_amount:,.2f}")
        self.logger.info(f"   - 💡 说明: 负数金额（退款/调整）已包含在净总金额计算中")
        
        # 按项目代码聚合
        if all_records:
            self.logger.info(f"\n🔄 开始聚合项目数据...")
            df_projects = pd.DataFrame(all_records)
            self.logger.debug(f"   - 待聚合记录数: {len(df_projects)}")
            
            grouped = df_projects.groupby(['project_code', 'project_name']).agg({
                'amount': 'sum',
                'person': 'count'
            }).reset_index()
            
            # 按成本降序排序
            grouped = grouped.sort_values('amount', ascending=False).reset_index(drop=True)
            
            total_count = len(grouped)
            self.logger.info(f"   ✅ 聚合完成，共 {total_count} 个唯一项目")
            
            # 验证金额总和
            grouped_total = grouped['amount'].sum()
            if abs(grouped_total - total_amount) > 0.01:
                self.logger.error(f"   ⚠️  金额验证失败！")
                self.logger.error(f"      原始总计: ¥{total_amount:,.2f}")
                self.logger.error(f"      聚合总计: ¥{grouped_total:,.2f}")
            
            self.logger.info(f"\n🏆 项目成本排名（Top {min(20, total_count)}）:")

            # 日志始终只显示前20个项目的详细信息（保持日志可读性）
            log_top_n = min(20, total_count)

            # 如果项目数量超过 top_n，将超出部分汇总到"其他"
            if total_count > top_n:
                self.logger.info(f"   - 展示前{top_n}个项目")
                self.logger.info(f"   - 其余{total_count - top_n}个项目汇总到\"其他\"")

                # 前 top_n 个项目（添加到结果）
                for idx, row in grouped.head(top_n).iterrows():
                    project_details = df_projects[
                        df_projects['project_code'] == row['project_code']
                    ].to_dict('records')

                    # 计算分类成本
                    project_df = df_projects[df_projects['project_code'] == row['project_code']]
                    flight_cost = project_df[project_df['type'] == '机票']['amount'].sum()
                    hotel_cost = project_df[project_df['type'] == '酒店']['amount'].sum()
                    train_cost = project_df[project_df['type'] == '火车票']['amount'].sum()

                    # 日志只输出前20个
                    if idx < log_top_n:
                        self.logger.info(f"\n   #{idx+1}. {row['project_code']} - {row['project_name']}")
                        self.logger.info(f"      总成本: ¥{row['amount']:,.2f} | 订单数: {int(row['person'])}")
                        self.logger.info(f"      ├─ 机票: ¥{flight_cost:,.2f}")
                        self.logger.info(f"      ├─ 酒店: ¥{hotel_cost:,.2f}")
                        self.logger.info(f"      └─ 火车票: ¥{train_cost:,.2f}")

                    results.append({
                        'project_code': row['project_code'],
                        'project_name': row['project_name'],
                        'total_cost': float(row['amount']),
                        'flight_cost': float(flight_cost),
                        'hotel_cost': float(hotel_cost),
                        'train_cost': float(train_cost),
                        'record_count': int(row['person']),
                        'details': project_details[:10]
                    })
                
                # 汇总"其他"项目
                others_df = grouped.iloc[top_n:]
                others_total_cost = float(others_df['amount'].sum())
                others_record_count = int(others_df['person'].sum())
                others_flight_cost = float(df_projects[df_projects['project_code'].isin(others_df['project_code']) & (df_projects['type'] == '机票')]['amount'].sum())
                others_hotel_cost = float(df_projects[df_projects['project_code'].isin(others_df['project_code']) & (df_projects['type'] == '酒店')]['amount'].sum())
                others_train_cost = float(df_projects[df_projects['project_code'].isin(others_df['project_code']) & (df_projects['type'] == '火车票')]['amount'].sum())
                
                self.logger.info(f"\n   #{top_n+1}. 其他")
                self.logger.info(f"      汇总项目数: {total_count - top_n}")
                self.logger.info(f"      总成本: ¥{others_total_cost:,.2f} | 订单数: {others_record_count}")
                
                results.append({
                    'project_code': '其他',
                    'project_name': f'其他项目（{total_count - top_n}个）',
                    'total_cost': others_total_cost,
                    'flight_cost': others_flight_cost,
                    'hotel_cost': others_hotel_cost,
                    'train_cost': others_train_cost,
                    'record_count': others_record_count,
                    'details': []
                })
            else:
                # 如果不超过 top_n，返回全部
                self.logger.info(f"   - 项目总数不超过{top_n}，返回全部")
                
                for idx, row in grouped.iterrows():
                    project_details = df_projects[
                        df_projects['project_code'] == row['project_code']
                    ].to_dict('records')
                    
                    # 计算分类成本
                    project_df = df_projects[df_projects['project_code'] == row['project_code']]
                    flight_cost = project_df[project_df['type'] == '机票']['amount'].sum()
                    hotel_cost = project_df[project_df['type'] == '酒店']['amount'].sum()
                    train_cost = project_df[project_df['type'] == '火车票']['amount'].sum()
                    
                    self.logger.info(f"\n   #{idx+1}. {row['project_code']} - {row['project_name']}")
                    self.logger.info(f"      总成本: ¥{row['amount']:,.2f} | 订单数: {int(row['person'])}")
                    self.logger.info(f"      ├─ 机票: ¥{flight_cost:,.2f}")
                    self.logger.info(f"      ├─ 酒店: ¥{hotel_cost:,.2f}")
                    self.logger.info(f"      └─ 火车票: ¥{train_cost:,.2f}")
                    
                    results.append({
                        'project_code': row['project_code'],
                        'project_name': row['project_name'],
                        'total_cost': float(row['amount']),
                        'flight_cost': float(flight_cost),
                        'hotel_cost': float(hotel_cost),
                        'train_cost': float(train_cost),
                        'record_count': int(row['person']),
                        'details': project_details[:10]
                    })
            
            # 最终汇总
            self.logger.info(f"\n" + "=" * 80)
            self.logger.info(f"✅ 项目成本归集完成")
            self.logger.info(f"=" * 80)
            self.logger.info(f"📊 最终统计:")
            self.logger.info(f"   - 返回项目数: {len(results)}")
            self.logger.info(f"   - 总成本: ¥{sum(r['total_cost'] for r in results):,.2f}")
            self.logger.info(f"   - 总订单数: {sum(r['record_count'] for r in results)}")
            self.logger.info("=" * 80 + "\n")
        else:
            self.logger.warning("⚠️  没有找到任何项目记录")

        return results, total_count
    
    def cross_check_attendance_travel(self) -> List[Dict[str, Any]]:
        """
        交叉验证：考勤数据 vs 差旅数据

        异常定义：考勤状态精确为"上班"（在办公室工作），但同一天有差旅消费（出差在外）
        - "上班" + 有差旅消费 = 异常（时间和地点冲突）
        - "公休日上班" + 有差旅消费 = 正常（周末加班出差）
        - "出差" + 有差旅消费 = 正常（出差状态）
        """
        anomalies = []

        # 获取考勤数据
        attendance_df = self.clean_attendance_data()
        if attendance_df.empty or '当日状态判断' not in attendance_df.columns:
            return anomalies
        if '日期' not in attendance_df.columns:
            return anomalies

        # 仅保留有日期的数据，提前计算日期字段，避免后续重复转换
        attendance_df = attendance_df.dropna(subset=['日期']).copy()
        if attendance_df.empty:
            return anomalies

        attendance_df['日期'] = attendance_df['日期'].dt.date
        attendance_df['当日状态判断'] = attendance_df['当日状态判断'].astype(str)
        if '一级部门' in attendance_df.columns:
            attendance_df['一级部门'] = attendance_df['一级部门'].fillna('未知部门')
        else:
            attendance_df['一级部门'] = '未知部门'

        # 只关注考勤状态精确为"上班"的记录（排除"公休日上班"、"出差"等）
        # 真正的异常是：在办公室上班，但同一天有差旅消费
        work_attendance = attendance_df[
            attendance_df['当日状态判断'] == '上班'
        ]
        if work_attendance.empty:
            return anomalies

        # 聚合所有差旅数据（姓名 + 消费日期 + 差旅类型），并缓存
        travel_df = self._get_combined_travel_df()
        if travel_df.empty:
            return anomalies

        travel_grouped = (
            travel_df.groupby(['姓名', '消费日期'])['差旅类型']
            .apply(list)
            .reset_index()
            .rename(columns={'消费日期': '日期'})
        )

        # 基于姓名+日期一次性关联，找出上班但有差旅消费的记录
        merged = work_attendance.merge(
            travel_grouped,
            on=['姓名', '日期'],
            how='inner'
        )

        for _, row in merged.iterrows():
            date_val = row.get('日期')
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)
            travel_list = row.get('差旅类型', []) or []
            name = row.get('姓名', '')
            anomalies.append({
                'name': name,
                'date': date_str,
                'department': row.get('一级部门', '未知部门'),
                'anomaly_type': 'A',
                'attendance_status': row.get('当日状态判断', ''),
                'travel_records': travel_list,
                'description': f'{name} 在 {date_str} 考勤显示上班（在办公室），但有 {",".join(travel_list)} 消费记录（出差在外），存在时间和地点冲突'
            })

        self.logger.info(f"交叉验证完成，发现 {len(anomalies)} 条异常记录（上班状态有差旅消费）")
        return anomalies
    
    def analyze_booking_behavior(self) -> Dict[str, Any]:
        """
        预订行为分析（机票）
        """
        df = self.clean_travel_data('机票')
        if df.empty or '提前预定天数' not in df.columns:
            return {}
        
        # 过滤有效数据
        valid_df = df[df['提前预定天数'].notna() & (df['提前预定天数'] >= 0)]
        
        if valid_df.empty:
            return {}
        
        amount_col = '授信金额' if '授信金额' in valid_df.columns else '金额'
        
        # 计算统计指标
        avg_advance = float(valid_df['提前预定天数'].mean())
        
        # 相关性分析
        correlation = float(valid_df[['提前预定天数', amount_col]].corr().iloc[0, 1])
        
        # 提前天数分布
        advance_distribution = valid_df['提前预定天数'].value_counts().to_dict()
        advance_distribution = {str(int(k)): int(v) for k, v in advance_distribution.items()}
        
        # 按提前天数分组的平均成本
        cost_by_advance = valid_df.groupby('提前预定天数')[amount_col].mean().reset_index()
        cost_by_advance_list = [
            {'advance_days': int(row['提前预定天数']), 'avg_cost': float(row[amount_col])}
            for _, row in cost_by_advance.iterrows()
        ]
        
        return {
            'avg_advance_days': round(avg_advance, 2),
            'correlation_advance_cost': round(correlation, 3),
            'advance_day_distribution': advance_distribution,
            'cost_by_advance_days': sorted(cost_by_advance_list, key=lambda x: x['advance_days'])
        }

    def count_over_standard_orders(self) -> Dict[str, Any]:
        """
        统计各差旅类型的超标订单数量

        业务规则：
        - 机票：超标类型包含“超折扣”或“超时间”
        - 酒店：按“是否超标”为“是”
        - 火车票：按“是否超标”为“是”
        """
        flight_df = self.clean_travel_data('机票')
        hotel_df = self.clean_travel_data('酒店')
        train_df = self.clean_travel_data('火车票')

        def _count_yes(df: pd.DataFrame, column: str) -> int:
            if df.empty or column not in df.columns:
                return 0
            return int(df[df[column].astype(str).str.contains('是', na=False)].shape[0])

        def _extract_over_types(value: str) -> List[str]:
            """
            提取机票的超标类型标签
            - 支持以空格、逗号、分号、斜杠等分隔的多标签格式
            - 若字符串中包含已知关键字（超折扣/超时间）但未分隔，也能捕获
            """
            if not value or pd.isna(value):
                return []

            raw = str(value)
            tokens = []

            # 常见分隔符拆分
            for part in re.split(r'[;,，、/\\s]+', raw):
                cleaned = part.strip()
                if cleaned and '超' in cleaned:
                    tokens.append(cleaned)

            # 兜底：处理未显式分隔但包含关键字的场景
            for keyword in ['超折扣', '超时间']:
                if keyword in raw and keyword not in tokens:
                    tokens.append(keyword)

            return tokens

        flight_over = 0
        flight_over_type_counter: Counter[str] = Counter()
        if not flight_df.empty:
            if '超标类型' in flight_df.columns:
                # 统计数量
                flight_over = int(
                    flight_df[
                        flight_df['超标类型']
                        .astype(str)
                        .str.contains('超折扣|超时间', na=False)
                    ].shape[0]
                )

                # 统计类型分布
                type_series = flight_df['超标类型'].dropna().astype(str)
                for raw_type in type_series:
                    for token in _extract_over_types(raw_type):
                        flight_over_type_counter[token] += 1

                # 如果有类型分布但未匹配到数量，用分布求和兜底
                if flight_over == 0 and flight_over_type_counter:
                    flight_over = sum(flight_over_type_counter.values())
            elif '是否超标' in flight_df.columns:
                flight_over = _count_yes(flight_df, '是否超标')
                if flight_over > 0:
                    flight_over_type_counter['未注明类型'] += flight_over

        hotel_over = _count_yes(hotel_df, '是否超标')
        train_over = _count_yes(train_df, '是否超标')

        total = flight_over + hotel_over + train_over

        return {
            'total': int(total),
            'flight': int(flight_over),
            'hotel': int(hotel_over),
            'train': int(train_over),
            'flight_over_types': {k: int(v) for k, v in flight_over_type_counter.items()},
        }

    def count_total_orders(self) -> Dict[str, int]:
        """
        统计各差旅类型及总订单数
        """
        flight_df = self.clean_travel_data('机票')
        hotel_df = self.clean_travel_data('酒店')
        train_df = self.clean_travel_data('火车票')

        flight = 0 if flight_df is None else int(len(flight_df))
        hotel = 0 if hotel_df is None else int(len(hotel_df))
        train = 0 if train_df is None else int(len(train_df))

        return {
            'total': int(flight + hotel + train),
            'flight': flight,
            'hotel': hotel,
            'train': train,
        }
    
    def calculate_department_costs(self, top_n: int = 15) -> List[Dict[str, Any]]:
        """
        部门成本汇总（包含平均工时和人数统计）
        
        Args:
            top_n: 返回前N个部门，其余汇总到"其他"（默认15）
        """
        results = []
        
        # 获取考勤数据以计算工时和人数
        attendance_df = self.clean_attendance_data()
        dept_attendance_stats = {}
        
        if not attendance_df.empty and '一级部门' in attendance_df.columns:
            # 计算每个部门的平均工时和人数
            for dept in attendance_df['一级部门'].unique():
                if pd.isna(dept):
                    continue
                dept_data = attendance_df[attendance_df['一级部门'] == dept]
                avg_hours = 0
                if '工时' in dept_data.columns:
                    valid_hours = dept_data[dept_data['工时'] != 0]['工时'].dropna()
                    if not valid_hours.empty:
                        avg_hours = float(valid_hours.mean())
                    if pd.isna(avg_hours):
                        avg_hours = 0
                person_count = dept_data['姓名'].nunique() if '姓名' in dept_data.columns else 0
                dept_attendance_stats[dept] = {
                    'avg_hours': float(avg_hours),
                    'person_count': int(person_count)
                }
        
        # 始终从明细表计算部门成本（不使用"差旅汇总" sheet）
        travel_data = {
            '机票': 'flight_cost',
            '酒店': 'hotel_cost',
            '火车票': 'train_cost'
        }
        
        dept_costs = {}
        
        for sheet_name, cost_key in travel_data.items():
            df = self.clean_travel_data(sheet_name)
            if df.empty:
                continue
            
            # 尝试关联部门信息（优先使用差旅表中的部门，如果没有则从考勤表获取）
            df = df.copy()
            if '一级部门' in df.columns:
                # 差旅表已有部门信息，优先使用
                pass
            elif not attendance_df.empty and '姓名' in attendance_df.columns and '一级部门' in attendance_df.columns:
                # 从考勤表获取部门信息
                name_dept = attendance_df[['姓名', '一级部门']].drop_duplicates()
                df = df.merge(name_dept, on='姓名', how='left')

            if '一级部门' not in df.columns:
                continue

            amount_col = '授信金额' if '授信金额' in df.columns else '金额'

            for _, row in df.iterrows():
                dept = row.get('一级部门')
                # 处理部门为空的情况
                if pd.isna(dept) or (isinstance(dept, str) and dept.strip() == ''):
                    dept = '未知部门'
                else:
                    dept = str(dept).strip()
                
                if dept not in dept_costs:
                    stats = dept_attendance_stats.get(dept, {'avg_hours': 0, 'person_count': 0})
                    dept_costs[dept] = {
                        'department': dept,
                        'total_cost': 0,
                        'flight_cost': 0,
                        'hotel_cost': 0,
                        'train_cost': 0,
                        'avg_hours': stats['avg_hours'],
                        'person_count': stats['person_count']
                    }
                
                amount = row.get(amount_col, 0) or 0
                dept_costs[dept][cost_key] += amount
                dept_costs[dept]['total_cost'] += amount
        
        results = list(dept_costs.values())
        results = sorted(results, key=lambda x: x['total_cost'], reverse=True)
        
        # 应用 top_n 限制并添加"其他"
        return self._apply_top_n_with_others(results, top_n, 'department')
    
    def _apply_top_n_with_others(self, results: List[Dict[str, Any]], top_n: int, name_key: str) -> List[Dict[str, Any]]:
        """
        应用 Top N 限制，将超出部分汇总到"其他"
        
        Args:
            results: 已排序的结果列表
            top_n: 保留的前N条记录
            name_key: 名称字段（'department' 或 'project_code'）
        
        Returns:
            处理后的结果列表
        """
        if not results or len(results) <= top_n:
            return results
        
        # 前 top_n 条
        top_results = results[:top_n]
        
        # 剩余的汇总到"其他"
        others = results[top_n:]
        total_count = len(results)
        
        others_summary = {
            name_key: '其他',
            'total_cost': sum(item.get('total_cost', 0) for item in others),
            'flight_cost': sum(item.get('flight_cost', 0) for item in others),
            'hotel_cost': sum(item.get('hotel_cost', 0) for item in others),
            'train_cost': sum(item.get('train_cost', 0) for item in others),
        }
        
        # 如果是部门数据，计算平均工时和总人数
        if name_key == 'department':
            avg_hours_list = [item.get('avg_hours', 0) for item in others if item.get('avg_hours', 0) > 0]
            others_summary['avg_hours'] = sum(avg_hours_list) / len(avg_hours_list) if avg_hours_list else 0
            others_summary['person_count'] = sum(item.get('person_count', 0) for item in others)
        
        # 如果是项目数据
        if name_key == 'project_code':
            others_summary['project_name'] = f'其他项目（{total_count - top_n}个）'
            others_summary['record_count'] = sum(item.get('record_count', 0) for item in others)
            others_summary['details'] = []
        
        top_results.append(others_summary)
        
        return top_results
    
    def get_attendance_summary(self) -> Dict[str, Any]:
        """
        考勤数据汇总
        """
        df = self.clean_attendance_data()
        if df.empty:
            return {}
        
        total_records = len(df)
        total_persons = df['姓名'].nunique() if '姓名' in df.columns else 0
        
        status_distribution = {}
        if '当日状态判断' in df.columns:
            status_distribution = df['当日状态判断'].value_counts().to_dict()
        
        avg_work_hours = 0
        if '工时' in df.columns:
            valid_hours = df[df['工时'] != 0]['工时'].dropna()
            if not valid_hours.empty:
                avg_work_hours = float(valid_hours.mean())
        
        return {
            'total_records': total_records,
            'total_persons': total_persons,
            'status_distribution': status_distribution,
            'avg_work_hours': round(avg_work_hours, 2)
        }
    
    def write_analysis_results(self, results: Dict[str, Any], output_path: Optional[str] = None) -> str:
        """
        将分析结果回写到 Excel（新增 Sheet）
        使用 openpyxl 保留原格式
        """
        if output_path is None:
            base_name = os.path.splitext(self.file_path)[0]
            output_path = f"{base_name}_analyzed.xlsx"
        
        if self.workbook is None:
            self.workbook = load_workbook(self.file_path)
        
        # 创建分析结果 Sheet
        sheet_name = "分析结果"
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        
        ws = self.workbook.create_sheet(sheet_name)
        
        # 写入标题
        ws.append(["CostMatrix 数据分析报告"])
        ws.append([f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])
        
        # 写入项目成本
        if 'project_costs' in results and results['project_costs']:
            ws.append(["项目成本归集"])
            ws.append(["项目代码", "项目名称", "总成本", "记录数"])
            for item in results['project_costs']:
                ws.append([
                    item['project_code'],
                    item['project_name'],
                    item['total_cost'],
                    item['record_count']
                ])
            ws.append([])
        
        # 写入部门成本
        if 'department_costs' in results and results['department_costs']:
            ws.append(["部门成本汇总"])
            ws.append(["部门", "总成本", "机票", "酒店", "火车票"])
            for item in results['department_costs']:
                ws.append([
                    item['department'],
                    item['total_cost'],
                    item['flight_cost'],
                    item['hotel_cost'],
                    item['train_cost']
                ])
            ws.append([])
        
        # 写入异常记录
        if 'anomalies' in results and results['anomalies']:
            ws.append(["交叉验证异常"])
            ws.append(["姓名", "日期", "异常类型", "考勤状态", "说明"])
            for item in results['anomalies']:
                ws.append([
                    item['name'],
                    item['date'],
                    item['anomaly_type'],
                    item['attendance_status'],
                    item['description']
                ])
        
        # 保存文件
        self.workbook.save(output_path)

        return output_path

    def get_all_project_details(self) -> List[Dict[str, Any]]:
        """
        获取所有项目的详细信息（包括人员、日期范围、超标等）

        Returns:
            包含所有项目详细信息的列表
        """
        self.logger.info("=" * 80)
        self.logger.info("开始获取所有项目详细信息")
        self.logger.info("=" * 80)

        results = []
        travel_sheets = ['机票', '酒店', '火车票']
        all_records = []

        # 收集所有差旅记录
        for sheet_name in travel_sheets:
            df = self.clean_travel_data(sheet_name)
            if df.empty or '项目' not in df.columns:
                continue

            amount_col = '授信金额' if '授信金额' in df.columns else '金额'
            date_cols = ['出发日期', '出发日期.1', '出发时间', '起飞日期', '起飞日期.1', '起飞时间', '起飞时间.1', '入住日期', '入住时间']
            date_col = next((col for col in date_cols if col in df.columns), None)

            # 获取考勤数据用于部门信息（作为备用）
            attendance_df = self.clean_attendance_data()
            person_dept_map = {}
            if not attendance_df.empty and '姓名' in attendance_df.columns and '一级部门' in attendance_df.columns:
                person_dept_map = attendance_df[['姓名', '一级部门']].drop_duplicates().set_index('姓名')['一级部门'].to_dict()

            for idx, row in df.iterrows():
                project_str = row.get('项目', '')
                project_code, project_name = self.extract_project_code(project_str)
                amount = row.get(amount_col, 0)
                person = row.get('姓名', '')
                date_val = row.get(date_col, '')

                # 优先使用差旅表中的部门信息，如果没有则从考勤表中查找
                department = None
                if '一级部门' in df.columns:
                    department = row.get('一级部门')
                    # 处理空值或NaN
                    if pd.isna(department) or (isinstance(department, str) and department.strip() == ''):
                        department = None
                if not department:
                    department = person_dept_map.get(person, '未知部门')
                else:
                    department = str(department).strip()

                # 处理日期
                if pd.notna(date_val):
                    if hasattr(date_val, 'strftime'):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val)
                else:
                    date_str = ''

                # 检查是否超标（需要正确判断字符串"是"或"否"）
                is_over_standard = False
                over_type = ''
                over_standard_val = row.get('是否超标', '')
                if pd.notna(over_standard_val):
                    is_over_standard = str(over_standard_val).strip() == '是'
                    if is_over_standard and '超标类型' in df.columns:
                        over_type = row.get('超标类型', '')

                # 计算提前预订天数
                advance_days = None
                if '预订日期' in df.columns and '出发日期' in df.columns:
                    book_date = row.get('预订日期')
                    dep_date = row.get('出发日期')
                    if pd.notna(book_date) and pd.notna(dep_date):
                        try:
                            if hasattr(book_date, 'to_pydatetime'):
                                book_date = book_date.to_pydatetime()
                            if hasattr(dep_date, 'to_pydatetime'):
                                dep_date = dep_date.to_pydatetime()
                            advance_days = (dep_date - book_date).days
                        except:
                            pass

                # 空项目处理
                if not project_code:
                    project_code = '空项目'
                    project_name = '未分配项目'

                all_records.append({
                    'project_code': project_code,
                    'project_name': project_name,
                    'person': person,
                    'department': department,
                    'type': sheet_name,
                    'amount': amount,
                    'date': date_str,
                    'is_over_standard': bool(is_over_standard),
                    'over_type': over_type,
                    'advance_days': advance_days
                })

        if not all_records:
            self.logger.warning("没有找到任何差旅记录")
            return []

        # 转换为 DataFrame
        df_all = pd.DataFrame(all_records)

        # 按项目分组统计
        grouped = df_all.groupby(['project_code', 'project_name']).agg({
            'amount': 'sum',
            'person': lambda x: list(set(x)),  # 去重的人员列表
            'department': lambda x: list(set(x)),  # 去重的部门列表
            'date': ['min', 'max'],  # 最早和最晚日期
            'type': 'count',  # 总订单数
            'is_over_standard': 'sum'  # 超标订单数
        }).reset_index()

        # 展平列名
        grouped.columns = ['project_code', 'project_name', 'total_cost', 'person_list',
                          'department_list', 'date_start', 'date_end', 'record_count', 'over_standard_count']

        # 计算各类型成本和订单数
        for sheet_name in travel_sheets:
            type_df = df_all[df_all['type'] == sheet_name]
            type_grouped = type_df.groupby(['project_code', 'project_name']).agg({
                'amount': 'sum',
                'type': 'count'
            }).reset_index()
            type_grouped.columns = ['project_code', 'project_name', f'{sheet_name}_cost', f'{sheet_name}_count']
            grouped = grouped.merge(type_grouped, on=['project_code', 'project_name'], how='left')

        # 填充空值
        for sheet_name in travel_sheets:
            grouped[f'{sheet_name}_cost'] = grouped[f'{sheet_name}_cost'].fillna(0)
            grouped[f'{sheet_name}_count'] = grouped[f'{sheet_name}_count'].fillna(0)

        # 按成本降序排序
        grouped = grouped.sort_values('total_cost', ascending=False).reset_index(drop=True)

        # 构建结果
        for _, row in grouped.iterrows():
            person_list = row['person_list'] if isinstance(row['person_list'], list) else []
            department_list = row['department_list'] if isinstance(row['department_list'], list) else []

            # 格式化日期
            date_start = row['date_start'] if pd.notna(row['date_start']) else ''
            date_end = row['date_end'] if pd.notna(row['date_end']) else ''

            results.append({
                'code': row['project_code'],
                'name': row['project_name'],
                'total_cost': float(row['total_cost']),
                'flight_cost': float(row.get('机票_cost', 0)),
                'hotel_cost': float(row.get('酒店_cost', 0)),
                'train_cost': float(row.get('火车票_cost', 0)),
                'record_count': int(row['record_count']),
                'flight_count': int(row.get('机票_count', 0)),
                'hotel_count': int(row.get('酒店_count', 0)),
                'train_count': int(row.get('火车票_count', 0)),
                'person_count': len(person_list),
                'person_list': person_list,
                'department_list': department_list,
                'date_range': {
                    'start': str(date_start),
                    'end': str(date_end)
                },
                'over_standard_count': int(row['over_standard_count'])
            })

        self.logger.info(f"✅ 共获取 {len(results)} 个项目的详细信息")
        self.logger.info("=" * 80 + "\n")

        return results

    def get_project_order_records(self, project_code: str) -> List[Dict[str, Any]]:
        """
        获取指定项目的所有订单记录

        Args:
            project_code: 项目代码

        Returns:
            该项目的所有订单记录列表
        """
        travel_sheets = ['机票', '酒店', '火车票']
        records = []

        # 获取考勤数据用于部门信息
        attendance_df = self.clean_attendance_data()
        person_dept_map = {}
        if not attendance_df.empty and '姓名' in attendance_df.columns and '一级部门' in attendance_df.columns:
            person_dept_map = attendance_df[['姓名', '一级部门']].drop_duplicates().set_index('姓名')['一级部门'].to_dict()

        for sheet_name in travel_sheets:
            df = self.clean_travel_data(sheet_name)
            if df.empty or '项目' not in df.columns:
                continue

            amount_col = '授信金额' if '授信金额' in df.columns else '金额'
            date_cols = ['出发日期', '出发日期.1', '出发时间', '起飞日期', '起飞日期.1', '起飞时间', '起飞时间.1', '入住日期', '入住时间']
            date_col = next((col for col in date_cols if col in df.columns), None)

            for idx, row in df.iterrows():
                project_str = row.get('项目', '')
                extracted_code, extracted_name = self.extract_project_code(project_str)

                # 空项目处理
                if not extracted_code:
                    extracted_code = '空项目'
                    extracted_name = '未分配项目'

                # 匹配项目代码
                if extracted_code != project_code:
                    continue

                amount = row.get(amount_col, 0)
                person = row.get('姓名', '')
                date_val = row.get(date_col, '')

                # 优先使用差旅表中的部门信息，如果没有则从考勤表中查找
                department = None
                if '一级部门' in df.columns:
                    department = row.get('一级部门')
                    # 处理空值或NaN
                    if pd.isna(department) or (isinstance(department, str) and department.strip() == ''):
                        department = None
                if not department:
                    department = person_dept_map.get(person, '未知部门')
                else:
                    department = str(department).strip()

                # 处理日期
                if pd.notna(date_val):
                    if hasattr(date_val, 'strftime'):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val)
                else:
                    date_str = ''

                # 检查是否超标（需要正确判断字符串"是"或"否"）
                is_over_standard = False
                over_type = ''
                over_standard_val = row.get('是否超标', '')
                if pd.notna(over_standard_val):
                    is_over_standard = str(over_standard_val).strip() == '是'
                    if is_over_standard and '超标类型' in df.columns:
                        over_type = row.get('超标类型', '')

                # 计算提前预订天数
                advance_days = None
                if '预订日期' in df.columns and '出发日期' in df.columns:
                    book_date = row.get('预订日期')
                    dep_date = row.get('出发日期')
                    if pd.notna(book_date) and pd.notna(dep_date):
                        try:
                            if hasattr(book_date, 'to_pydatetime'):
                                book_date = book_date.to_pydatetime()
                            if hasattr(dep_date, 'to_pydatetime'):
                                dep_date = dep_date.to_pydatetime()
                            advance_days = (dep_date - book_date).days
                        except:
                            pass

                # 转换类型名称
                type_mapping = {'机票': 'flight', '酒店': 'hotel', '火车票': 'train'}

                records.append({
                    'id': f"{sheet_name}_{idx}",
                    'project_code': extracted_code,
                    'project_name': extracted_name,
                    'person': person,
                    'department': department,
                    'type': type_mapping.get(sheet_name, 'other'),
                    'amount': float(amount),
                    'date': date_str,
                    'is_over_standard': bool(is_over_standard),
                    'over_type': over_type,
                    'advance_days': advance_days
                })

        # 按日期排序
        records.sort(key=lambda x: x['date'], reverse=True)

        return records

    def get_department_hierarchy(self) -> Dict[str, Any]:
        """
        获取部门层级结构

        Returns:
            {
                'level1': ['一级部门1', '一级部门2', ...],
                'level2': {'一级部门1': ['二级部门1', '二级部门2'], ...},
                'level3': {'二级部门1': ['三级部门1', '三级部门2'], ...}
            }
        """
        df = self.clean_attendance_data()
        if df.empty:
            return {'level1': [], 'level2': {}, 'level3': {}}

        result = {
            'level1': [],
            'level2': {},
            'level3': {}
        }

        # 获取一级部门
        if '一级部门' in df.columns:
            result['level1'] = sorted(df['一级部门'].dropna().unique().tolist())

        # 获取二级部门（按一级部门分组）
        if '一级部门' in df.columns and '二级部门' in df.columns:
            for l1 in result['level1']:
                l2_list = df[df['一级部门'] == l1]['二级部门'].dropna().unique().tolist()
                result['level2'][l1] = sorted(l2_list)

        # 获取三级部门（按二级部门分组）
        if '二级部门' in df.columns and '三级部门' in df.columns:
            for l1, l2_list in result['level2'].items():
                for l2 in l2_list:
                    l3_list = df[df['二级部门'] == l2]['三级部门'].dropna().unique().tolist()
                    result['level3'][l2] = sorted(l3_list)

        return result

    def get_department_list(self, level: int, parent: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        获取部门列表

        Args:
            level: 部门层级 (1=一级, 2=二级, 3=三级)
            parent: 父部门名称（level>1时必需）

        Returns:
            部门列表，每个部门包含人数、成本、平均工时等信息
        """
        df = self.clean_attendance_data()
        if df.empty:
            self.logger.warning(f"get_department_list: 考勤数据为空，level={level}, parent={parent}")
            return []

        # 确定部门列名（中文数字：一级部门、二级部门、三级部门）
        level_name_map = {1: '一级部门', 2: '二级部门', 3: '三级部门'}
        dept_col = level_name_map.get(level)

        if not dept_col or dept_col not in df.columns:
            self.logger.warning(f"get_department_list: 缺少部门列 '{dept_col}'（level={level}），可用列: {df.columns.tolist()}")
            return []

        # 筛选部门
        filtered_df = df.copy()

        if level == 2 and parent:
            filtered_df = filtered_df[filtered_df['一级部门'] == parent]
        elif level == 3 and parent:
            filtered_df = filtered_df[filtered_df['二级部门'] == parent]

        # 检查筛选后的数据
        if filtered_df.empty:
            self.logger.warning(f"get_department_list: 筛选后数据为空，level={level}, parent={parent}")
            return []

        departments = filtered_df[dept_col].dropna().unique().tolist()
        self.logger.info(f"get_department_list: 找到 {len(departments)} 个{dept_col}: {departments[:5]}...")

        # 获取差旅数据用于成本计算
        dept_costs = self._calculate_costs_by_department(filtered_df, dept_col)

        results = []
        for dept in departments:
            dept_data = filtered_df[filtered_df[dept_col] == dept]

            # 计算人数
            person_count = dept_data['姓名'].nunique() if '姓名' in dept_data.columns else 0

            # 计算平均工时
            avg_hours = 0
            if '工时' in dept_data.columns:
                valid_hours = dept_data[dept_data['工时'] != 0]['工时'].dropna()
                if not valid_hours.empty:
                    avg_hours = float(valid_hours.mean())

            # 获取成本
            cost_info = dept_costs.get(dept, {'total_cost': 0, 'flight_cost': 0, 'hotel_cost': 0, 'train_cost': 0})

            results.append({
                'name': dept,
                'level': level,
                'parent': parent,
                'person_count': int(person_count),
                'total_cost': float(cost_info['total_cost']),
                'avg_work_hours': round(avg_hours, 2)
            })

        # 按成本降序排序
        results.sort(key=lambda x: x['total_cost'], reverse=True)

        self.logger.info(f"get_department_list: 返回 {len(results)} 个部门，前3个: {[(r['name'], r['total_cost'], r['person_count']) for r in results[:3]]}")

        return results

    def _calculate_costs_by_department(self, attendance_df: pd.DataFrame, dept_col: str) -> Dict[str, Dict[str, float]]:
        """
        计算各部门的差旅成本

        Args:
            attendance_df: 考勤数据（已按部门筛选）
            dept_col: 部门列名

        Returns:
            {部门名: {total_cost, flight_cost, hotel_cost, train_cost}}
        """
        dept_costs = {}

        if attendance_df.empty:
            self.logger.warning(f"_calculate_costs_by_department: attendance_df 为空")
            return dept_costs

        # 获取姓名到部门的映射（保留所有映射，一个人可能属于多个部门）
        name_dept_map = {}
        for _, row in attendance_df[['姓名', dept_col]].drop_duplicates().iterrows():
            name = row.get('姓名')
            dept = row.get(dept_col)
            if not name or pd.isna(name) or not dept or pd.isna(dept):
                continue
            if name not in name_dept_map:
                name_dept_map[name] = []
            if dept not in name_dept_map[name]:
                name_dept_map[name].append(dept)

        # 遍历差旅数据
        travel_sheets = ['机票', '酒店', '火车票']
        cost_keys = {'机票': 'flight_cost', '酒店': 'hotel_cost', '火车票': 'train_cost'}

        total_records = 0
        matched_records = 0

        for sheet_name in travel_sheets:
            df = self.clean_travel_data(sheet_name)
            if df.empty:
                continue

            amount_col = '授信金额' if '授信金额' in df.columns else '金额'

            for _, row in df.iterrows():
                total_records += 1
                name = row.get('姓名', '')
                if not name or pd.isna(name):
                    continue

                depts = name_dept_map.get(name, [])
                if not depts:
                    continue

                matched_records += 1
                amount = row.get(amount_col, 0) or 0
                cost_key = cost_keys[sheet_name]

                # 一个人可能属于多个部门，将成本分配到所有关联部门
                for dept in depts:
                    if dept not in dept_costs:
                        dept_costs[dept] = {'total_cost': 0, 'flight_cost': 0, 'hotel_cost': 0, 'train_cost': 0}
                    dept_costs[dept][cost_key] += amount
                    dept_costs[dept]['total_cost'] += amount

        self.logger.info(f"_calculate_costs_by_department: 差旅记录 {total_records} 条，匹配 {matched_records} 条，部门数 {len(dept_costs)}")
        return dept_costs

    def get_department_detail_metrics(self, department_name: str, level: int = 3) -> Dict[str, Any]:
        """
        获取指定部门的详细指标（12项）

        Args:
            department_name: 部门名称
            level: 部门层级 (1=一级, 2=二级, 3=三级)

        Returns:
            包含12项指标的字典
        """
        df = self.clean_attendance_data()
        if df.empty:
            return {}

        # 确定部门列名（使用中文数字：一级部门、二级部门、三级部门）
        level_name_map = {1: '一级部门', 2: '二级部门', 3: '三级部门'}
        dept_col = level_name_map.get(level)
        if not dept_col or dept_col not in df.columns:
            return {}

        # 筛选该部门的数据
        dept_df = df[df[dept_col] == department_name].copy()

        if dept_df.empty:
            return {}

        # 获取父部门
        parent_dept = None
        if level == 2 and '一级部门' in dept_df.columns:
            parent_dept = dept_df['一级部门'].dropna().unique()
            parent_dept = parent_dept[0] if len(parent_dept) > 0 else None
        elif level == 3 and '二级部门' in dept_df.columns:
            parent_dept = dept_df['二级部门'].dropna().unique()
            parent_dept = parent_dept[0] if len(parent_dept) > 0 else None

        # 1. 当月考勤天数分布
        attendance_days_distribution = {}
        if '当日状态判断' in dept_df.columns:
            attendance_days_distribution = dept_df['当日状态判断'].value_counts().to_dict()

        # 2. 公休日上班天数
        weekend_work_days = 0
        if '当日状态判断' in dept_df.columns:
            weekend_work_days = int(dept_df[dept_df['当日状态判断'] == '公休日上班'].shape[0])

        # 3. 工作日出勤天数
        workday_attendance_days = 0
        if '当日状态判断' in dept_df.columns:
            workday_attendance_days = int(dept_df[dept_df['当日状态判断'] == '上班'].shape[0])

        # 4. 工作日平均工时
        avg_work_hours = 0
        if '工时' in dept_df.columns:
            valid_hours = dept_df[(dept_df['当日状态判断'] == '上班') & (dept_df['工时'] != 0)]['工时'].dropna()
            if not valid_hours.empty:
                avg_work_hours = float(valid_hours.mean())

        # 5. 出差天数
        travel_days = 0
        if '当日状态判断' in dept_df.columns:
            travel_days = int(dept_df[dept_df['当日状态判断'] == '出差'].shape[0])

        # 6. 请假天数
        leave_days = 0
        if '当日状态判断' in dept_df.columns:
            leave_days = int(dept_df[dept_df['当日状态判断'] == '请假'].shape[0])

        # 7. 异常天数（通过交叉验证）
        anomalies = self.cross_check_attendance_travel()
        dept_anomalies = [a for a in anomalies if a.get('department') == department_name]
        anomaly_days = len(dept_anomalies)

        # 8. 晚上7:30后下班人数
        late_after_1930_count = 0
        if '最晚19:30之后' in dept_df.columns:
            late_after_1930_count = int(dept_df[dept_df['最晚19:30之后'] == '符合']['姓名'].nunique())

        # 9. 周末出勤次数
        weekend_attendance_count = 0
        if '日期' in dept_df.columns and '当日状态判断' in dept_df.columns:
            dept_df['weekday'] = dept_df['日期'].dt.dayofweek
            weekend_df = dept_df[dept_df['weekday'].isin([5, 6])]  # 5=周六, 6=周日
            weekend_attendance_count = int(weekend_df[weekend_df['当日状态判断'].isin(['上班', '出差'])].shape[0])

        # 10. 出差排行榜（按出差天数）
        travel_ranking = []
        if '当日状态判断' in dept_df.columns:
            travel_df = dept_df[dept_df['当日状态判断'] == '出差']
            if not travel_df.empty and '姓名' in travel_df.columns:
                travel_counts = travel_df['姓名'].value_counts().head(10)
                travel_ranking = [
                    {'name': name, 'value': int(count), 'detail': f'{count}天'}
                    for name, count in travel_counts.items()
                ]

        # 11. 异常排行榜（按异常次数）
        anomaly_ranking = []
        if dept_anomalies:
            from collections import Counter
            anomaly_counts = Counter([a.get('name', '') for a in dept_anomalies])
            anomaly_ranking = [
                {'name': name, 'value': int(count), 'detail': f'{count}次'}
                for name, count in anomaly_counts.most_common(10)
            ]

        # 12. 最晚下班排行榜
        latest_checkout_ranking = []
        if '最晚打卡时间' in dept_df.columns:
            dept_df['punch_time'] = pd.to_datetime(dept_df['最晚打卡时间'], format='%H:%M:%S', errors='coerce')
            valid_punch = dept_df[dept_df['punch_time'].notna()].sort_values('punch_time', ascending=False)
            if not valid_punch.empty and '姓名' in valid_punch.columns:
                for _, row in valid_punch.head(10).iterrows():
                    latest_checkout_ranking.append({
                        'name': row['姓名'],
                        'value': 0,  # ECharts需要数值，这里仅用于排序
                        'detail': row['最晚打卡时间']
                    })

        # 13. 最长工时排行榜（按平均工时排名）
        longest_hours_ranking = []
        if '工时' in dept_df.columns and '姓名' in dept_df.columns:
            # 先按人员分组计算平均工时，排除工时为0的记录
            person_avg_hours = dept_df[dept_df['工时'].notna() & (dept_df['工时'] != 0)].groupby('姓名')['工时'].mean()
            # 按平均工时降序排列
            person_avg_hours = person_avg_hours.sort_values(ascending=False)
            # 取前10名
            for name, avg_hours in person_avg_hours.head(10).items():
                longest_hours_ranking.append({
                    'name': name,
                    'value': float(round(avg_hours, 2)),
                    'detail': f'{avg_hours:.2f}小时'
                })

        return {
            'department_name': department_name,
            'department_level': f'{level}级部门',
            'parent_department': parent_dept,
            'attendance_days_distribution': attendance_days_distribution,
            'weekend_work_days': weekend_work_days,
            'workday_attendance_days': workday_attendance_days,
            'avg_work_hours': round(avg_work_hours, 2),
            'travel_days': travel_days,
            'leave_days': leave_days,
            'anomaly_days': anomaly_days,
            'late_after_1930_count': late_after_1930_count,
            'weekend_attendance_count': weekend_attendance_count,
            'travel_ranking': travel_ranking,
            'anomaly_ranking': anomaly_ranking,
            'latest_checkout_ranking': latest_checkout_ranking,
            'longest_hours_ranking': longest_hours_ranking
        }

    def get_level1_department_statistics(self, level1_name: str) -> Dict[str, Any]:
        """
        获取一级部门的汇总统计数据（用于二级部门表格下方的统计展示）

        Args:
            level1_name: 一级部门名称

        Returns:
            包含以下统计数据的字典:
            - total_travel_cost: 累计差旅成本
            - attendance_days_distribution: 考勤天数分布
            - travel_ranking: 出差排行榜（按人）
            - avg_hours_ranking: 平均工时排行榜（按人）
            - level2_department_stats: 二级部门统计列表（包含所有指标）
        """
        df = self.clean_attendance_data()
        if df.empty:
            return {}

        # 筛选该一级部门的数据
        level1_df = df[df['一级部门'] == level1_name].copy()

        if level1_df.empty:
            return {}

        # 1. 累计差旅成本
        total_travel_cost = 0
        dept_costs = self._calculate_costs_by_department(level1_df, '二级部门')
        for cost_info in dept_costs.values():
            total_travel_cost += cost_info['total_cost']

        # 2. 考勤天数分布（整个一级部门）
        attendance_days_distribution = {}
        if '当日状态判断' in level1_df.columns:
            attendance_days_distribution = level1_df['当日状态判断'].value_counts().to_dict()

        # 3. 出差排行榜（按人，在整个一级部门范围内）
        travel_ranking = []
        if '当日状态判断' in level1_df.columns:
            travel_df = level1_df[level1_df['当日状态判断'] == '出差']
            if not travel_df.empty and '姓名' in travel_df.columns:
                travel_counts = travel_df['姓名'].value_counts().head(10)
                travel_ranking = [
                    {'name': name, 'value': int(count), 'detail': f'{count}天'}
                    for name, count in travel_counts.items()
                ]

        # 4. 平均工时排行榜（按人，在整个一级部门范围内）
        avg_hours_ranking = []
        if '工时' in level1_df.columns and '姓名' in level1_df.columns:
            person_avg_hours = level1_df[level1_df['工时'].notna() & (level1_df['工时'] != 0)].groupby('姓名')['工时'].mean()
            person_avg_hours = person_avg_hours.sort_values(ascending=False)
            for name, avg_hours in person_avg_hours.head(10).items():
                avg_hours_ranking.append({
                    'name': name,
                    'value': float(round(avg_hours, 2)),
                    'detail': f'{avg_hours:.2f}小时'
                })

        # 5. 二级部门统计（包含所有指标）
        level2_department_stats = []
        if '二级部门' in level1_df.columns:
            level2_list = level1_df['二级部门'].dropna().unique().tolist()

            for l2_dept in level2_list:
                l2_df = level1_df[level1_df['二级部门'] == l2_dept]

                # 计算人数
                person_count = l2_df['姓名'].nunique() if '姓名' in l2_df.columns else 0

                # 计算平均工时
                avg_hours = 0
                if '工时' in l2_df.columns:
                    valid_hours = l2_df[(l2_df['当日状态判断'] == '上班') & (l2_df['工时'] != 0)]['工时'].dropna()
                    if not valid_hours.empty:
                        avg_hours = float(valid_hours.mean())

                # 工作日出勤天数
                workday_attendance_days = 0
                if '当日状态判断' in l2_df.columns:
                    workday_attendance_days = int(l2_df[l2_df['当日状态判断'] == '上班'].shape[0])

                # 公休日上班天数
                weekend_work_days = 0
                if '当日状态判断' in l2_df.columns:
                    weekend_work_days = int(l2_df[l2_df['当日状态判断'] == '公休日上班'].shape[0])

                # 周末出勤次数
                weekend_attendance_count = 0
                if '日期' in l2_df.columns and '当日状态判断' in l2_df.columns:
                    l2_df_copy = l2_df.copy()
                    l2_df_copy['weekday'] = l2_df_copy['日期'].dt.dayofweek
                    weekend_df = l2_df_copy[l2_df_copy['weekday'].isin([5, 6])]
                    weekend_attendance_count = int(weekend_df[weekend_df['当日状态判断'].isin(['上班', '出差'])].shape[0])

                # 出差天数
                travel_days = 0
                if '当日状态判断' in l2_df.columns:
                    travel_days = int(l2_df[l2_df['当日状态判断'] == '出差'].shape[0])

                # 请假天数
                leave_days = 0
                if '当日状态判断' in l2_df.columns:
                    leave_days = int(l2_df[l2_df['当日状态判断'] == '请假'].shape[0])

                # 异常天数
                anomalies = self.cross_check_attendance_travel()
                dept_anomalies = [a for a in anomalies if a.get('department') == l2_dept]
                anomaly_days = len(dept_anomalies)

                # 晚上7:30后下班人数
                late_after_1930_count = 0
                if '最晚19:30之后' in l2_df.columns:
                    late_after_1930_count = int(l2_df[l2_df['最晚19:30之后'] == '符合']['姓名'].nunique())

                # 获取该二级部门的成本
                cost_info = dept_costs.get(l2_dept, {'total_cost': 0})

                level2_department_stats.append({
                    'name': l2_dept,
                    'person_count': person_count,
                    'avg_work_hours': round(avg_hours, 2),
                    'workday_attendance_days': workday_attendance_days,
                    'weekend_work_days': weekend_work_days,
                    'weekend_attendance_count': weekend_attendance_count,
                    'travel_days': travel_days,
                    'leave_days': leave_days,
                    'anomaly_days': anomaly_days,
                    'late_after_1930_count': late_after_1930_count,
                    'total_cost': float(cost_info['total_cost'])
                })

            # 按成本降序排序
            level2_department_stats.sort(key=lambda x: x['total_cost'], reverse=True)

        return {
            'department_name': level1_name,
            'total_travel_cost': round(total_travel_cost, 2),
            'attendance_days_distribution': attendance_days_distribution,
            'travel_ranking': travel_ranking,
            'avg_hours_ranking': avg_hours_ranking,
            'level2_department_stats': level2_department_stats
        }

    def get_available_months(self) -> List[str]:
        """获取所有可用的月份列表（从差旅数据中提取，格式：YYYY-M，按时间升序排列）"""
        # Ensure data is loaded
        if not self.sheets_data:
            self.load_all_sheets()

        months_set = set()

        flight_df = self.clean_travel_data('机票')
        if not flight_df.empty and '起飞日期' in flight_df.columns:
            months = flight_df['起飞日期'].dt.strftime('%Y-%m').dropna().unique()
            months_set.update(months)

        hotel_df = self.clean_travel_data('酒店')
        if not hotel_df.empty and '入住日期' in hotel_df.columns:
            months = hotel_df['入住日期'].dt.strftime('%Y-%m').dropna().unique()
            months_set.update(months)

        train_df = self.clean_travel_data('火车票')
        if not train_df.empty and '出发日期' in train_df.columns:
            months = train_df['出发日期'].dt.strftime('%Y-%m').dropna().unique()
            months_set.update(months)

        return sorted(list(months_set))

    def _save_cache(self, cache_path: str, data: Dict[str, Any]):
        """Save analysis results to JSON cache file"""
        import json
        from pathlib import Path

        cache_file = Path(cache_path)
        cache_file.parent.mkdir(parents=True, exist_ok=True)

        temp_path = cache_file.with_suffix('.tmp')
        try:
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            temp_path.replace(cache_file)
            self.logger.info(f"Cache saved to: {cache_path}")
        except Exception as e:
            self.logger.error(f"Failed to save cache: {e}")
            if temp_path.exists():
                temp_path.unlink()
            raise
