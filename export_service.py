"""
Excel å¯¼å‡ºæœåŠ¡æ¨¡å—
ä½¿ç”¨ openpyxl åœ¨åŸ Excel æ–‡ä»¶ä¸­è¿½åŠ åˆ†æç»“æœ Sheet
ä¿æŒåŸæ–‡ä»¶çš„æ ·å¼å’Œæ’ç‰ˆ
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from typing import Dict, List
from io import BytesIO
import os
from logger_config import get_logger


class ExcelExporter:
    """Excel å¯¼å‡ºæœåŠ¡"""
    
    def __init__(self, file_path: str):
        """
        åˆå§‹åŒ–å¯¼å‡ºæœåŠ¡
        
        Args:
            file_path: åŸå§‹ Excel æ–‡ä»¶è·¯å¾„
        """
        self.file_path = file_path
        self.workbook = None
        self.logger = get_logger("export_service")
        
        self.logger.info(f"åˆå§‹åŒ–Excelå¯¼å‡ºæœåŠ¡ï¼Œæ–‡ä»¶è·¯å¾„: {file_path}")
    
    def load_workbook(self):
        """åŠ è½½åŸå§‹ Excel å·¥ä½œç°¿"""
        try:
            self.logger.debug("å¼€å§‹åŠ è½½Excelå·¥ä½œç°¿")
            self.workbook = openpyxl.load_workbook(self.file_path)
            sheet_count = len(self.workbook.sheetnames)
            self.logger.info(f"Excelå·¥ä½œç°¿åŠ è½½æˆåŠŸï¼Œå…± {sheet_count} ä¸ªå·¥ä½œè¡¨")
            self.logger.debug(f"å·¥ä½œè¡¨åˆ—è¡¨: {', '.join(self.workbook.sheetnames)}")
        except Exception as e:
            self.logger.error(f"æ— æ³•åŠ è½½ Excel æ–‡ä»¶: {str(e)}", exc_info=True)
            raise ValueError(f"æ— æ³•åŠ è½½ Excel æ–‡ä»¶: {str(e)}")
    
    def add_dashboard_sheet(self, dashboard_data: Dict):
        """
        æ·»åŠ  Dashboard_Data Sheet
        
        Args:
            dashboard_data: åˆ†æç»“æœå­—å…¸
        """
        self.logger.debug("å¼€å§‹æ·»åŠ Dashboard_Dataå·¥ä½œè¡¨")
        
        if self.workbook is None:
            self.load_workbook()
        
        # åˆ é™¤å·²å­˜åœ¨çš„åŒå Sheet
        if "Dashboard_Data" in self.workbook.sheetnames:
            self.logger.debug("åˆ é™¤å·²å­˜åœ¨çš„Dashboard_Dataå·¥ä½œè¡¨")
            del self.workbook["Dashboard_Data"]
        
        # åˆ›å»ºæ–° Sheet
        ws = self.workbook.create_sheet("Dashboard_Data", 0)
        self.logger.debug("Dashboard_Dataå·¥ä½œè¡¨åˆ›å»ºæˆåŠŸ")
        
        # è®¾ç½®æ ‡é¢˜æ ·å¼
        title_font = Font(name='å¾®è½¯é›…é»‘', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # è®¾ç½®å†…å®¹æ ·å¼
        content_font = Font(name='å¾®è½¯é›…é»‘', size=11)
        content_alignment = Alignment(horizontal='left', vertical='center')
        
        # è¾¹æ¡†æ ·å¼
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_idx = 1
        
        # 1. KPI æ±‡æ€»
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws[f'A{row_idx}']
        cell.value = 'ğŸ“Š KPI æ ¸å¿ƒæŒ‡æ ‡'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row_idx += 1
        
        kpi = dashboard_data.get('kpi', {})
        kpi_data = [
            ['æŒ‡æ ‡åç§°', 'æ•°å€¼', '', 'æŒ‡æ ‡åç§°', 'æ•°å€¼', ''],
            ['æ€»å·®æ—…æˆæœ¬', f"Â¥{kpi.get('total_cost', 0):,.2f}", '', 'æ€»è®¢å•æ•°', kpi.get('total_orders', 0), ''],
            ['å¼‚å¸¸è®°å½•æ•°', kpi.get('anomaly_count', 0), '', 'è¶…æ ‡è®¢å•æ•°', kpi.get('over_standard_count', 0), ''],
            ['ç´§æ€¥é¢„è®¢æ¯”ä¾‹', f"{kpi.get('urgent_booking_ratio', 0)}%", '', '', '', '']
        ]
        
        for row_data in kpi_data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
            row_idx += 1
        
        row_idx += 2
        
        # 2. é¡¹ç›®æˆæœ¬ Top 10
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws[f'A{row_idx}']
        cell.value = 'ğŸ’° é¡¹ç›®æˆæœ¬ Top 10'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row_idx += 1
        
        # è¡¨å¤´
        project_headers = ['é¡¹ç›®ä»£ç ', 'æ€»æˆæœ¬', 'æœºç¥¨æˆæœ¬', 'é…’åº—æˆæœ¬', 'ç«è½¦ç¥¨æˆæœ¬', 'è®¢å•æ•°é‡']
        for col_idx, header in enumerate(project_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(name='å¾®è½¯é›…é»‘', size=11, bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = title_alignment
            cell.border = thin_border
        row_idx += 1
        
        # æ•°æ®è¡Œ
        top_projects = dashboard_data.get('top_projects', [])
        for project in top_projects:
            row_data = [
                project.get('é¡¹ç›®ä»£ç ', ''),
                project.get('æ€»æˆæœ¬', 0),
                project.get('æœºç¥¨æˆæœ¬', 0),
                project.get('é…’åº—æˆæœ¬', 0),
                project.get('ç«è½¦ç¥¨æˆæœ¬', 0),
                project.get('è®¢å•æ•°é‡', 0)
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
            row_idx += 1
        
        row_idx += 2
        
        # 3. éƒ¨é—¨æŒ‡æ ‡
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws[f'A{row_idx}']
        cell.value = 'ğŸ¢ éƒ¨é—¨æŒ‡æ ‡æ±‡æ€»'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row_idx += 1
        
        # è¡¨å¤´
        dept_headers = ['ä¸€çº§éƒ¨é—¨', 'æ€»æˆæœ¬', 'æ€»å·¥æ—¶', 'äººå‘˜æ•°é‡', 'é¥±å’Œåº¦(%)', '']
        for col_idx, header in enumerate(dept_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(name='å¾®è½¯é›…é»‘', size=11, bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = title_alignment
            cell.border = thin_border
        row_idx += 1
        
        # æ•°æ®è¡Œ
        dept_metrics = dashboard_data.get('department_metrics', [])
        for dept in dept_metrics:
            row_data = [
                dept.get('ä¸€çº§éƒ¨é—¨', ''),
                dept.get('æ€»æˆæœ¬', 0),
                dept.get('æ€»å·¥æ—¶', 0),
                dept.get('äººå‘˜æ•°é‡', 0),
                dept.get('é¥±å’Œåº¦', 0),
                ''
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
            row_idx += 1
        
        # è°ƒæ•´åˆ—å®½
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        kpi = dashboard_data.get('kpi', {})
        top_projects_count = len(dashboard_data.get('top_projects', []))
        dept_metrics_count = len(dashboard_data.get('department_metrics', []))
        
        self.logger.info(f"Dashboard_Dataå·¥ä½œè¡¨æ·»åŠ å®Œæˆ: "
                        f"KPIæŒ‡æ ‡å·²æ·»åŠ , "
                        f"Topé¡¹ç›®={top_projects_count}, "
                        f"éƒ¨é—¨æŒ‡æ ‡={dept_metrics_count}")
    
    def add_anomaly_sheet(self, anomalies: List[Dict]):
        """
        æ·»åŠ  Anomaly_Log Sheet
        
        Args:
            anomalies: å¼‚å¸¸è®°å½•åˆ—è¡¨
        """
        self.logger.debug(f"å¼€å§‹æ·»åŠ Anomaly_Logå·¥ä½œè¡¨ï¼Œå¼‚å¸¸è®°å½•æ•°: {len(anomalies)}")
        
        if self.workbook is None:
            self.load_workbook()
        
        # åˆ é™¤å·²å­˜åœ¨çš„åŒå Sheet
        if "Anomaly_Log" in self.workbook.sheetnames:
            self.logger.debug("åˆ é™¤å·²å­˜åœ¨çš„Anomaly_Logå·¥ä½œè¡¨")
            del self.workbook["Anomaly_Log"]
        
        # åˆ›å»ºæ–° Sheet
        ws = self.workbook.create_sheet("Anomaly_Log")
        self.logger.debug("Anomaly_Logå·¥ä½œè¡¨åˆ›å»ºæˆåŠŸ")
        
        # è®¾ç½®æ ·å¼
        header_font = Font(name='å¾®è½¯é›…é»‘', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        content_font = Font(name='å¾®è½¯é›…é»‘', size=10)
        content_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # è¡¨å¤´
        headers = ['å¼‚å¸¸ç±»å‹', 'å§“å', 'æ—¥æœŸ', 'è€ƒå‹¤çŠ¶æ€', 'å·®æ—…ç±»å‹', 'å·®æ—…é‡‘é¢', 'ä¸€çº§éƒ¨é—¨', 'æè¿°']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # æ•°æ®è¡Œ
        for row_idx, anomaly in enumerate(anomalies, start=2):
            row_data = [
                anomaly.get('Type', ''),
                anomaly.get('å§“å', ''),
                anomaly.get('æ—¥æœŸ', ''),
                anomaly.get('è€ƒå‹¤çŠ¶æ€', ''),
                anomaly.get('å·®æ—…ç±»å‹', ''),
                anomaly.get('å·®æ—…é‡‘é¢', 0),
                anomaly.get('ä¸€çº§éƒ¨é—¨', ''),
                anomaly.get('æè¿°', '')
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
                
                # æ ¹æ®å¼‚å¸¸ç±»å‹è®¾ç½®èƒŒæ™¯è‰²
                if col_idx == 1:
                    if anomaly.get('Type') == 'Conflict':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    elif anomaly.get('Type') == 'NoExpense':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # è°ƒæ•´åˆ—å®½
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 40
        
        # å†»ç»“é¦–è¡Œ
        ws.freeze_panes = 'A2'
        
        conflict_count = len([a for a in anomalies if a.get('Type') == 'Conflict'])
        no_expense_count = len([a for a in anomalies if a.get('Type') == 'NoExpense'])
        
        self.logger.info(f"Anomaly_Logå·¥ä½œè¡¨æ·»åŠ å®Œæˆ: "
                        f"æ€»å¼‚å¸¸æ•°={len(anomalies)}, "
                        f"å†²çªç±»å‹={conflict_count}, "
                        f"æ— æ¶ˆè´¹ç±»å‹={no_expense_count}")
    
    def save_to_bytes(self) -> BytesIO:
        """
        ä¿å­˜å·¥ä½œç°¿åˆ°å†…å­˜å­—èŠ‚æµ
        
        Returns:
            BytesIO å¯¹è±¡
        """
        if self.workbook is None:
            self.logger.error("å·¥ä½œç°¿æœªåŠ è½½ï¼Œæ— æ³•ä¿å­˜")
            raise ValueError("å·¥ä½œç°¿æœªåŠ è½½")
        
        self.logger.debug("å¼€å§‹å°†å·¥ä½œç°¿ä¿å­˜åˆ°å†…å­˜å­—èŠ‚æµ")
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        size_bytes = output.getbuffer().nbytes
        self.logger.info(f"å·¥ä½œç°¿å·²ä¿å­˜åˆ°å†…å­˜ï¼Œå¤§å°: {size_bytes} bytes ({size_bytes/1024:.2f} KB)")
        return output
    
    def save_to_file(self, output_path: str):
        """
        ä¿å­˜å·¥ä½œç°¿åˆ°æ–‡ä»¶
        
        Args:
            output_path: è¾“å‡ºæ–‡ä»¶è·¯å¾„
        """
        if self.workbook is None:
            raise ValueError("å·¥ä½œç°¿æœªåŠ è½½")
        
        self.workbook.save(output_path)
    
    def export_with_analysis(
        self, 
        dashboard_data: Dict, 
        anomalies: List[Dict]
    ) -> BytesIO:
        """
        å¯¼å‡ºåŒ…å«åˆ†æç»“æœçš„ Excel æ–‡ä»¶
        
        Args:
            dashboard_data: Dashboard æ•°æ®
            anomalies: å¼‚å¸¸è®°å½•åˆ—è¡¨
            
        Returns:
            BytesIO å¯¹è±¡
        """
        self.logger.info("å¼€å§‹å¯¼å‡ºåŒ…å«åˆ†æç»“æœçš„Excelæ–‡ä»¶")
        
        self.load_workbook()
        self.add_dashboard_sheet(dashboard_data)
        self.add_anomaly_sheet(anomalies)
        result = self.save_to_bytes()
        
        self.logger.info("Excelæ–‡ä»¶å¯¼å‡ºå®Œæˆ")
        return result


